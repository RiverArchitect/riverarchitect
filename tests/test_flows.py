"""Tests for seasonal flow duration curves, annual peaks and return periods.

A synthetic record with a known distribution has a known duration curve, so these assert
against arithmetic rather than against recorded output.
"""

import datetime as dt

import numpy as np
import pytest

from riverarchitect.flows import (FlowSeries, GUMBEL_EULER, read_flow_series,
                                  return_periods, seasonal_flow_duration)


def record(start=dt.date(2019, 10, 1), days=366, flow=lambda i: 100.0 + i):
    return {start + dt.timedelta(days=i): flow(i) for i in range(days)}


# ------------------------------------------------------------------- reading

def test_read_flow_series_from_csv(tmp_path):
    path = tmp_path / "series.csv"
    path.write_text("Date,Mean daily\n2020-01-01,10\n2020-01-02,20\n", encoding="utf-8")
    series = read_flow_series(str(path))
    assert series == {dt.date(2020, 1, 1): 10.0, dt.date(2020, 1, 2): 20.0}


def test_read_flow_series_skips_the_unit_row(tmp_path):
    """The original's template carries a units row under the header."""
    path = tmp_path / "series.csv"
    path.write_text("Date,Mean daily\n(DD-MMM-YY),(CFS)\n2020-01-01,10\n", encoding="utf-8")
    assert read_flow_series(str(path)) == {dt.date(2020, 1, 1): 10.0}


def test_a_record_with_no_usable_row_is_rejected(tmp_path):
    path = tmp_path / "series.csv"
    path.write_text("Date,Mean daily\nnonsense,nonsense\n", encoding="utf-8")
    with pytest.raises(ValueError, match="no usable date"):
        read_flow_series(str(path))


# ------------------------------------------------------------ duration curve

def test_duration_curve_is_a_descending_rank_series():
    """Rank i of n is exceeded i/n*100 per cent - the original's plotting position."""
    flows = FlowSeries({dt.date(2020, 1, day): float(day) for day in range(1, 5)})
    discharge, exceedance = flows.duration_curve()
    assert list(discharge) == [4.0, 3.0, 2.0, 1.0]
    assert list(exceedance) == [25.0, 50.0, 75.0, 100.0]


def test_the_lowest_flow_is_exceeded_all_of_the_time():
    flows = FlowSeries(record())
    _discharge, exceedance = flows.duration_curve()
    assert exceedance[-1] == pytest.approx(100.0)
    assert 0.0 < exceedance[0] <= 1.0


def test_a_season_selects_only_its_own_days():
    """February flows say nothing about a June-to-October lifestage."""
    flows = FlowSeries(record())
    whole_year = flows.season_flows()
    winter = flows.season_flows((1, 1), (2, 28))
    assert winter.size == 31 + 28          # 29 February falls outside the bounds given
    assert flows.season_flows((1, 1), (2, 29)).size == 31 + 29   # 2020 is a leap year
    assert winter.size < whole_year.size


def test_a_season_may_wrap_the_turn_of_the_year():
    """1 October to 30 September is the hydrological year, not an empty range."""
    flows = FlowSeries(record())
    assert flows.season_flows((10, 1), (9, 30)).size == len(flows.series)
    # a genuinely short wrapping season keeps only its own days
    assert flows.season_flows((12, 30), (1, 2)).size == 4


def test_exceedance_is_interpolated_on_the_curve():
    flows = FlowSeries({dt.date(2020, 1, day): float(day) * 100 for day in range(1, 5)})
    resolved = flows.exceedance_of([150.0, 400.0])
    # between rank 4 (100 cfs, 100%) and rank 3 (200 cfs, 75%)
    assert resolved[150.0] == pytest.approx(87.5)
    assert resolved[400.0] == pytest.approx(25.0)


def test_exceedance_clamps_outside_the_record():
    flows = FlowSeries(record())
    resolved = flows.exceedance_of([1.0, 1e9])
    assert resolved[1.0] == pytest.approx(100.0)      # below everything on record
    assert resolved[1e9] == pytest.approx(0.0)        # above everything on record


def test_a_season_the_record_does_not_cover_is_rejected():
    flows = FlowSeries({dt.date(2020, 6, day): 10.0 for day in range(1, 10)})
    with pytest.raises(ValueError, match="no flow in the record"):
        flows.duration_curve((1, 1), (2, 1))


# ------------------------------------------------------------- annual peaks

def test_annual_peaks_use_the_water_year():
    series = record(dt.date(2019, 10, 1), 366)
    series[dt.date(2019, 12, 25)] = 5000.0        # water year 2020
    flows = FlowSeries(series)
    peaks = flows.annual_peaks()
    assert peaks[2020] == pytest.approx(5000.0)
    assert 2019 not in peaks or peaks[2019] < 5000.0


def test_calendar_years_are_available_too():
    series = record(dt.date(2019, 10, 1), 366)
    series[dt.date(2019, 12, 25)] = 5000.0
    peaks = FlowSeries(series).annual_peaks(water_year_start=(1, 1))
    assert peaks[2019] == pytest.approx(5000.0)


# ----------------------------------------------------------- return periods

def test_return_period_rises_with_discharge():
    peaks = {year: 1000.0 + 100.0 * ((year * 7) % 11) for year in range(1970, 2020)}
    periods = return_periods(peaks, [1000.0, 1500.0, 3000.0])
    assert periods[1000.0] < periods[1500.0] < periods[3000.0]
    assert all(value >= 1.0 for value in periods.values())


def test_the_gumbel_fit_matches_the_closed_form():
    peaks = [1000.0 + 137.0 * ((i * 13) % 17) for i in range(40)]
    values = np.asarray(sorted(peaks))
    scale = values.std(ddof=1) * np.sqrt(6.0) / np.pi
    location = values.mean() - GUMBEL_EULER * scale

    target = 2500.0
    expected = 1.0 / (1.0 - np.exp(-np.exp(-(target - location) / scale)))
    assert return_periods(peaks, [target])[target] == pytest.approx(expected, rel=1e-9)


def test_too_few_peaks_is_refused_rather_than_guessed():
    with pytest.raises(ValueError, match="at least 10"):
        return_periods([100.0, 200.0, 300.0])


# ------------------------------------------------------------------ writing

def test_written_workbook_is_the_layout_sharc_reads(tmp_path):
    from riverarchitect.sharc import read_flow_duration

    flows = FlowSeries(record(days=400))
    path = tmp_path / "flow_duration_test.xlsx"
    flows.write_flow_duration(str(path), discharges=[150.0, 300.0, 450.0], code="test")

    duration = read_flow_duration(str(path))
    assert set(duration) == {150.0, 300.0, 450.0}
    # a higher discharge is exceeded less of the time
    assert duration[150.0] > duration[300.0] > duration[450.0]


def test_a_long_record_is_sampled_rather_than_truncated(tmp_path):
    """Truncating would throw away either the floods or the low flows."""
    flows = FlowSeries(record(days=5000))
    path = tmp_path / "flow_duration_long.xlsx"
    flows.write_flow_duration(str(path), limit=100)

    import openpyxl
    sheet = openpyxl.load_workbook(str(path))["Q data"]
    written = [sheet.cell(row, 1).value for row in range(3, sheet.max_row + 1)]
    written = [value for value in written if value is not None]
    assert len(written) <= 100
    # both ends of the record survive the sampling
    assert max(written) == pytest.approx(max(flows.series.values()))
    assert min(written) == pytest.approx(min(flows.series.values()))


def test_seasonal_flow_duration_writes_one_workbook_per_lifestage(tmp_path, monkeypatch):
    pytest.importorskip("openpyxl")
    from riverarchitect import config

    monkeypatch.setenv("RIVERARCHITECT_HOME", str(tmp_path))
    config.set_project_home(str(tmp_path))
    try:
        written = seasonal_flow_duration(record(days=740), "demo",
                                         discharges=[100.0, 500.0, 900.0],
                                         output_dir=str(tmp_path / "flows"))
    finally:
        config.set_project_home(None)

    assert written, "no flow duration curve was built"
    codes = {entry["code"] for entry in written}
    assert "chsp" in codes and "chfr" in codes
    for entry in written:
        assert entry["path"].endswith("flow_duration_%s.xlsx" % entry["code"])
        assert entry["days_in_season"] > 0


def test_lifestages_with_different_seasons_get_different_curves(tmp_path):
    """The whole point of a *seasonal* curve: a winter species sees winter flows."""
    pytest.importorskip("openpyxl")

    # low all year, high only in February
    series = {}
    day = dt.date(2019, 10, 1)
    while day <= dt.date(2020, 9, 30):
        series[day] = 5000.0 if day.month == 2 else 100.0
        day += dt.timedelta(days=1)

    flows = FlowSeries(series)
    # Chinook fry: 1 Feb - 15 Jun. Chinook juvenile: 16 Jun - 29 Nov, no February at all.
    fry = flows.exceedance_of([1000.0], *flows.fish.season_dates("Chinook salmon", "fry"))
    juvenile = flows.exceedance_of([1000.0],
                                   *flows.fish.season_dates("Chinook salmon", "juvenile"))
    assert fry[1000.0] > juvenile[1000.0]
    assert juvenile[1000.0] == pytest.approx(0.0)
