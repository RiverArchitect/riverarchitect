"""Flow analysis: seasonal flow duration curves and annual peak series.

The open-source replacement for the *Analyze Flows* part of the ArcGIS ``GetStarted``
module (``riverpy/cFlows.py``) and for the ``Tools/make_annual_*.py`` scripts.

Why it matters
--------------
:mod:`riverarchitect.sharc` integrates usable habitat area over a **flow duration curve** to
produce SHArea, the single number a habitat project is judged on. That curve is not
something the hydrodynamic model produces: it comes from a gauged daily flow record, and it
is *seasonal* - Chinook fry occupy the reach between 1 February and 15 June, so the flows of
the rest of the year say nothing about their habitat. Without this module a project has 2D
results, habitat suitability curves and no way to weigh one discharge against another.

The chain
---------
1. Read a daily record of dates and mean daily discharge (:class:`FlowSeries`).
2. Take every flow that falls inside a lifestage's season, in every year of the record, and
   sort them descending. Rank *i* of *n* is exceeded ``i / n * 100`` per cent of that season
   - the seasonal flow duration curve (:meth:`FlowSeries.duration_curve`).
3. Interpolate that curve at each discharge the 2D model was run for
   (:meth:`FlowSeries.exceedance_of`), because those are the only discharges habitat can be
   evaluated at.
4. Write ``00_Flows/<condition>/flow_duration_<code>.xlsx`` in the layout
   :func:`riverarchitect.sharc.read_flow_duration` expects
   (:meth:`FlowSeries.write_flow_duration`).

Also here: :meth:`FlowSeries.annual_peaks`, the annual maximum series that a flood frequency
analysis needs to attach a return period to each modelled discharge - the
``input_definitions.inp`` "Return periods" line that lifespan mapping reads. The original
exported it for HEC-SSP; :func:`return_periods` additionally fits a Gumbel distribution so
that a first estimate needs no second program.

Relation to the original
------------------------
``cFlows.SeasonalFlowProcessor`` built a ``n_days x n_years`` matrix by walking the record
day by day, with a hand-rolled leap-day skip and an index that had to stay in step with the
file. The result is a plain "collect the flows inside the season, sort, rank", which is what
this does. Leap days are kept rather than dropped: excluding 29 February biased the curve by
one day in four years for no reason the original documented.
"""

import datetime as dt
import logging
import os

import numpy as np

from . import config

__all__ = ["FlowSeries", "read_flow_series", "seasonal_flow_duration", "return_periods",
           "GUMBEL_EULER"]

logger = logging.getLogger("riverarchitect")

#: Euler-Mascheroni constant, for the Gumbel moment estimator.
GUMBEL_EULER = 0.5772156649015329


def read_flow_series(path, date_column=0, discharge_column=1, sheet=None):
    """Read a daily flow record from a spreadsheet or CSV.

    Args:
        path (str): ``.xlsx``, ``.xls`` or ``.csv`` with a date column and a mean daily
            discharge column.
        date_column (int): zero-based index of the date column.
        discharge_column (int): zero-based index of the discharge column.
        sheet (str or int): worksheet, for a workbook with more than one.

    Returns:
        dict: ``{datetime.date: discharge}``, ascending by date.
    """
    import warnings

    import pandas as pd

    if str(path).lower().endswith((".xlsx", ".xls")):
        frame = pd.read_excel(path, sheet_name=0 if sheet is None else sheet)
    else:
        frame = pd.read_csv(path)

    with warnings.catch_warnings():
        # A gauge export routinely carries a units row under the header, and mixed content
        # in the date column is exactly what `errors="coerce"` is for. Pandas warning that
        # it fell back to dateutil is not something the caller can act on.
        warnings.simplefilter("ignore", UserWarning)
        dates = pd.to_datetime(frame.iloc[:, date_column], errors="coerce")
    values = pd.to_numeric(frame.iloc[:, discharge_column], errors="coerce")
    series = {}
    for date, value in zip(dates, values):
        if pd.isna(date) or pd.isna(value):
            continue
        series[date.date()] = float(value)
    if not series:
        raise ValueError("no usable date/discharge pair in %s - check the column indices"
                         % path)
    return dict(sorted(series.items()))


def _in_season(date, start, end):
    """Is ``date`` inside a season given as ``(month, day)`` bounds?

    A season whose end falls before its start wraps the turn of the year, which is how a
    hydrological year is written: 1 October to 30 September.
    """
    key = (date.month, date.day)
    if start <= end:
        return start <= key <= end
    return key >= start or key <= end


class FlowSeries:
    """A daily flow record, and the duration curves derived from it.

    Args:
        series (dict or str): ``{date: discharge}``, or a path to read one from.
        unit (str): ``"us"`` or ``"si"``; only used for labels.
        fish (riverarchitect.sharc.FishDatabase): the database seasons are read from.
            Built on demand.

    Attributes:
        series (dict): the record, ascending by date.
    """

    def __init__(self, series, unit="us", fish=None):
        self.series = read_flow_series(series) if isinstance(series, str) \
            else dict(sorted(series.items()))
        if not self.series:
            raise ValueError("the flow record is empty")
        self.unit = str(unit).lower()
        self._fish = fish
        self.logger = logger

    @property
    def fish(self):
        """The habitat database, loaded on first use."""
        if self._fish is None:
            from .sharc import FishDatabase
            self._fish = FishDatabase()
        return self._fish

    @property
    def years(self):
        """Calendar years the record spans."""
        return sorted({date.year for date in self.series})

    def __len__(self):
        return len(self.series)

    def __repr__(self):
        return "FlowSeries(%d days, %d-%d)" % (len(self.series), self.years[0],
                                               self.years[-1])

    # ------------------------------------------------------------------ duration

    def season_flows(self, start=None, end=None):
        """Every discharge inside a season, across every year of the record.

        Args:
            start, end (tuple): ``(month, day)`` bounds. Omit both for the whole year.

        Returns:
            numpy.ndarray: the flows, unsorted.
        """
        if start is None or end is None:
            values = list(self.series.values())
        else:
            values = [q for date, q in self.series.items() if _in_season(date, start, end)]
        return np.asarray([q for q in values if np.isfinite(q)], dtype="float64")

    def duration_curve(self, start=None, end=None):
        """Seasonal flow duration curve: discharge against per-cent exceedance.

        Rank *i* of *n* sorted descending is exceeded ``i / n * 100`` per cent of the
        season, which is the plotting position the original used. The highest flow therefore
        carries a small non-zero exceedance rather than zero, and the lowest carries exactly
        100.

        Returns:
            tuple: ``(discharge, exceedance)``, discharge descending.
        """
        flows = np.sort(self.season_flows(start, end))[::-1]
        if flows.size == 0:
            raise ValueError("no flow in the record falls inside the season given")
        exceedance = np.arange(1, flows.size + 1) / flows.size * 100.0
        return flows, exceedance

    def exceedance_of(self, discharges, start=None, end=None):
        """Per-cent of the season each discharge is exceeded.

        Interpolated on the duration curve. A discharge above everything on record is
        exceeded 0 per cent of the time and one below everything 100 per cent, which is what
        ``cFlows.interpolate_flow_exceedance`` returned at those ends.

        Args:
            discharges (iterable): the discharges to evaluate, normally the ones the 2D
                model was run for.
            start, end (tuple): season bounds, as ``(month, day)``.

        Returns:
            dict: ``{discharge: per-cent exceedance}``.
        """
        flows, exceedance = self.duration_curve(start, end)
        # numpy.interp needs ascending x; the curve is discharge-descending.
        order = np.argsort(flows)
        return {float(q): float(np.interp(float(q), flows[order], exceedance[order],
                                          left=100.0, right=0.0))
                for q in discharges}

    # --------------------------------------------------------------- annual peaks

    def annual_peaks(self, water_year_start=(10, 1)):
        """Largest discharge of each water year.

        Args:
            water_year_start (tuple): ``(month, day)`` the water year begins on. The
                default is 1 October, the U.S. convention. Pass ``(1, 1)`` for calendar
                years.

        Returns:
            dict: ``{water year: peak discharge}``. The year label is the one the water
            year *ends* in, so 1 October 2019 to 30 September 2020 is 2020.
        """
        peaks = {}
        month, day = water_year_start
        for date, discharge in self.series.items():
            year = date.year + 1 if (date.month, date.day) >= (month, day) else date.year
            if (month, day) == (1, 1):
                year = date.year
            if not np.isfinite(discharge):
                continue
            peaks[year] = max(peaks.get(year, float("-inf")), discharge)
        return dict(sorted(peaks.items()))

    # ------------------------------------------------------------------- writing

    def write_flow_duration(self, path, discharges=None, start=None, end=None,
                            code="", condition="", limit=20000):
        """Write a ``flow_duration_<code>.xlsx`` in the layout SHArC reads.

        Two sheets, matching the workbooks the original produced:

        ``Q data``
            the full duration curve, discharge in column A and per-cent exceedance in
            column B from row 3, with the species code and the season in columns D and E.
        ``2D data``
            the modelled discharges in column B and their interpolated exceedance in
            column C from row 3 - the sheet
            :func:`riverarchitect.sharc.read_flow_duration` reads.

        Args:
            path (str): where to write.
            discharges (iterable): the modelled discharges for the ``2D data`` sheet.
            start, end (tuple): season bounds as ``(month, day)``.
            code (str): four-letter species/lifestage code, for the header.
            condition (str): condition name, for the header.
            limit (int): cap on the number of ``Q data`` rows written. A 50-year record has
                18 000 of them; beyond a point they are only file size.

        Returns:
            str: the path written.
        """
        import openpyxl

        flows, exceedance = self.duration_curve(start, end)
        if flows.size > limit:
            # Keep the shape of the curve: sample evenly along it rather than truncating,
            # which would throw away either the floods or the low flows.
            keep = np.unique(np.linspace(0, flows.size - 1, limit).astype(int))
            flows, exceedance = flows[keep], exceedance[keep]

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Q data"
        sheet.cell(1, 1, "FLOW")
        sheet.cell(1, 2, "DATA")
        sheet.cell(2, 1, "AVERAGE Q")
        sheet.cell(2, 2, "% Exceedance")
        sheet.cell(4, 4, "Eco shortname:")
        sheet.cell(4, 5, code)
        sheet.cell(5, 4, "Season start date:")
        sheet.cell(5, 5, " Month:%d Day:%d" % start if start else " whole year")
        sheet.cell(6, 4, "end date:")
        sheet.cell(6, 5, " Month:%d Day:%d" % end if end else " whole year")
        sheet.cell(7, 4, "Record:")
        sheet.cell(7, 5, "%d-%d" % (self.years[0], self.years[-1]))
        sheet.cell(8, 4, "Condition:")
        sheet.cell(8, 5, condition)
        for row, (discharge, value) in enumerate(zip(flows, exceedance), start=3):
            sheet.cell(row, 1, float(discharge))
            sheet.cell(row, 2, float(value))

        two_d = workbook.create_sheet("2D data")
        two_d.cell(1, 2, "2D")
        two_d.cell(1, 3, "MODEL")
        two_d.cell(2, 2, "Q")
        two_d.cell(2, 3, "Intp. % Exceedance")
        if discharges:
            resolved = self.exceedance_of(discharges, start, end)
            # Descending discharge, so the sheet reads like the original's.
            for row, discharge in enumerate(sorted(resolved, reverse=True), start=3):
                two_d.cell(row, 2, float(discharge))
                two_d.cell(row, 3, float(resolved[discharge]))

        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        workbook.save(path)
        self.logger.info("   >> wrote %s (%d curve points, %d modelled discharges)",
                         path, flows.size, len(discharges or []))
        return path


def seasonal_flow_duration(series, condition, discharges=None, pairs=None, unit="us",
                           output_dir=None, fish=None):
    """Build a flow duration workbook for every species and lifestage.

    This is *Analyze Flows*: the step between a gauge record and a SHArea calculation. One
    workbook per species and lifestage lands in ``00_Flows/<condition>/``, named with the
    four-letter code SHArC looks for, so :meth:`riverarchitect.sharc.SHArC.run` finds it
    without being told where it is.

    Args:
        series (FlowSeries or dict or str): the daily record.
        condition (str): condition name; sets the output folder and the discharges.
        discharges (iterable): the modelled discharges. Defaults to every ``h<Q>.tif`` in
            the condition.
        pairs (iterable): ``(species, lifestage)`` tuples. Defaults to every pair the fish
            database defines that has a season.
        unit (str): ``"us"`` or ``"si"``.
        output_dir (str): where the workbooks go. Defaults to ``00_Flows/<condition>/``.
        fish (riverarchitect.sharc.FishDatabase): the database to read seasons from.

    Returns:
        list: one summary dict per workbook written.
    """
    from .condition import Condition

    flows = series if isinstance(series, FlowSeries) \
        else FlowSeries(series, unit=unit, fish=fish)

    if discharges is None:
        entry = Condition(condition)
        discharges = sorted({Condition.discharge_of(name)
                             for name in entry.all_depth_rasters()} - {None})
    discharges = sorted(float(q) for q in discharges)

    if pairs is None:
        pairs = list(flows.fish.pairs())

    output_dir = output_dir or os.path.join(config.dir_flows(), str(condition))
    os.makedirs(output_dir, exist_ok=True)

    written = []
    for species, lifestage in pairs:
        season = flows.fish.season_dates(species, lifestage)
        if season is None:
            logger.info("      * no season for %s %s - skipped", species, lifestage)
            continue
        start, end = season
        code = flows.fish.shortname(species, lifestage)
        path = os.path.join(output_dir, "flow_duration_%s.xlsx" % code)
        try:
            flows.write_flow_duration(path, discharges, start, end, code=code,
                                      condition=str(condition))
        except ValueError as exc:      # a season the record does not cover
            logger.info("      * %s %s: %s", species, lifestage, exc)
            continue
        exceedance = flows.exceedance_of(discharges, start, end)
        written.append({
            "species": species,
            "lifestage": lifestage,
            "code": code,
            "season": season,
            "path": path,
            "days_in_season": int(flows.season_flows(start, end).size),
            "exceedance": exceedance,
        })
    return written


def return_periods(peaks, discharges=None):
    """Flood return periods from an annual maximum series, by a Gumbel fit.

    The ``Return periods`` line of ``input_definitions.inp`` is what turns a set of modelled
    discharges into a lifespan axis, and it is the input most often missing. The original
    exported the annual peaks for HEC-SSP and left the fitting to it; this gives a defensible
    first estimate in one step, by the method of moments on a Gumbel (EV1) distribution:

    .. math::

        \\alpha = \\frac{s\\sqrt{6}}{\\pi}, \\quad u = \\bar{x} - \\gamma\\alpha, \\quad
        T = \\frac{1}{1 - \\exp(-\\exp(-(x - u) / \\alpha))}

    Args:
        peaks (dict or iterable): annual peak discharges, or ``{year: peak}``.
        discharges (iterable): discharges to evaluate. Defaults to the peaks themselves.

    Returns:
        dict: ``{discharge: return period in years}``, never below 1.0.

    Raises:
        ValueError: with fewer than ten peaks, where a Gumbel fit is not worth quoting.
    """
    values = np.asarray(sorted(peaks.values()) if isinstance(peaks, dict) else sorted(peaks),
                        dtype="float64")
    values = values[np.isfinite(values)]
    if values.size < 10:
        raise ValueError("a Gumbel fit needs at least 10 annual peaks, got %d"
                         % values.size)

    scale = values.std(ddof=1) * np.sqrt(6.0) / np.pi
    location = values.mean() - GUMBEL_EULER * scale
    if scale <= 0:
        raise ValueError("the annual peaks have no spread - cannot fit a distribution")

    result = {}
    for discharge in (values if discharges is None else discharges):
        discharge = float(discharge)
        non_exceedance = np.exp(-np.exp(-(discharge - location) / scale))
        # Clip just below 1 so a discharge far beyond the record gives a large finite
        # period rather than infinity.
        non_exceedance = min(non_exceedance, 1.0 - 1e-9)
        result[discharge] = max(1.0 / (1.0 - non_exceedance), 1.0)
    return result
