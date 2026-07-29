"""SHArC: habitat suitability and Seasonal Habitat Area.

The open-source replacement for the ArcGIS ``SHArC`` module. It turns 2D hydrodynamic
results into a map of how good the habitat is for a given fish species and lifestage, and
then into a single number a project can be judged on.

The chain
---------
1. **Habitat suitability curves** come from ``Fish.xlsx``: for each species and lifestage, a
   piecewise-linear curve mapping flow depth, flow velocity, substrate size or cover radius
   onto a suitability index between 0 and 1.
2. **Hydraulic HSI rasters** apply those curves to the depth and velocity rasters of a
   discharge, giving ``dsi`` (depth) and ``vsi`` (velocity).
3. **Composite HSI** (``cHSI``) combines them, by geometric mean or product, optionally
   including a cover HSI, and is masked to the wetted area.
4. **Usable area** at a discharge is the area where ``cHSI`` exceeds a threshold, optionally
   weighted by the mean ``cHSI`` there.
5. **SHArea** integrates usable area over the flow duration curve, so a habitat area that
   only exists at a rare discharge counts for little.

Relation to the original
------------------------
``cHSI.nested_con_raster_calc`` built one ``Con()`` raster per curve segment and took their
cell-wise maximum. That is a piecewise-linear interpolation written the only way arcpy map
algebra allowed; here it is :func:`apply_curve`, which is :func:`numpy.interp` with the same
end behaviour - hold the first value below the curve, drop to zero above it. The SHArea
integration reproduces the ``=(E5-E4)/100*F5`` column of the original
``CONDITION_sharea_template.xlsx``.
"""

import glob
import logging
import os

import numpy as np

from . import config, raster
from .condition import Condition

__all__ = ["FishDatabase", "apply_curve", "cover_hsi", "SHArC", "COMBINE_METHODS",
           "COVER_TYPES", "GRAIN_SIZE_LIMITS"]

logger = logging.getLogger("riverarchitect")

#: Ways of combining the individual suitability indices into a composite one.
COMBINE_METHODS = ("geometric_mean", "product")

#: Row of each habitat suitability curve in ``Fish.xlsx``, and of the scalar thresholds.
#: Reproduces ``cFish.Fish.parameter_rows`` of the original.
PARAMETER_ROWS = {"u": 9, "h": 38, "substrate": 72, "cobbles": 81, "boulders": 82,
                  "plants": 84, "wood": 85, "start_date": 6, "end_date": 7,
                  "h_min": 87, "u_max": 88}

#: Column offsets a lifestage can occupy within its species' block. The *label* at each
#: offset is read from the workbook rather than assumed, because it varies by species:
#: offset 3 is "fry" for salmon but "ammocoetes" for lamprey, and the "All Aquatic" column
#: block uses "hydrological year", "season", "depth > x" and "velocity > x" instead.
LIFESTAGE_OFFSETS = (1, 3, 5, 7)

_SPECIES_ROW = 2
_LIFESTAGE_ROW = 5
_FIRST_SPECIES_COLUMN = 3          # column C
_COLUMNS_PER_SPECIES = 8


def default_fish_database():
    """Path of the ``Fish.xlsx`` shipped with the package."""
    return os.path.join(config.templates_dir(), "Fish.xlsx")


class FishDatabase:
    """Habitat suitability curves, read from a ``Fish.xlsx`` in the original layout.

    Args:
        path (str): workbook path. Defaults to the one shipped with the package.

    Attributes:
        species (list): species common names, in workbook order.
    """

    def __init__(self, path=None):
        import warnings

        import openpyxl

        self.path = path or default_fish_database()
        if not os.path.isfile(self.path):
            raise FileNotFoundError("no fish database at %s" % self.path)
        with warnings.catch_warnings():
            # openpyxl warns that it drops the workbook's data-validation extension on
            # read. It does not affect the curve values, and there is nothing to act on.
            warnings.simplefilter("ignore", UserWarning)
            self._sheet = openpyxl.load_workbook(self.path, data_only=True).active

        self.species = []
        self._species_column = {}
        self._lifestages = {}          # species -> [label, ...] in workbook order
        self._offsets = {}             # species -> {label: column offset}
        self._scan()

    # ------------------------------------------------------------------ lookup

    def resolve_species(self, species):
        """Match a species name against the workbook, ignoring case and spacing.

        The workbook writes ``"Chinook Salmon"``; papers, file names and the rest of this
        package write ``"Chinook salmon"``. Matching exactly would make a caller's choice of
        capitalisation decide whether an analysis runs, so it does not.

        Args:
            species (str): the name to look up.

        Returns:
            str: the name as the workbook spells it.

        Raises:
            KeyError: when no species matches.
        """
        wanted = " ".join(str(species).split()).lower()
        for name in self.species:
            if " ".join(name.split()).lower() == wanted:
                return name
        raise KeyError("no such species in %s: %r (have: %s)"
                       % (self.path, species, ", ".join(self.species)))

    def resolve_lifestage(self, species, lifestage):
        """Match a lifestage of a species, ignoring case and spacing.

        Returns:
            str: the label as the workbook spells it.

        Raises:
            KeyError: when the species has no such lifestage.
        """
        species = self.resolve_species(species)
        wanted = " ".join(str(lifestage).split()).lower()
        for label in self._lifestages.get(species, []):
            if " ".join(label.split()).lower() == wanted:
                return label
        raise KeyError("species %r has no lifestage %r (have: %s)"
                       % (species, lifestage, ", ".join(self._lifestages.get(species, []))))

    def _scan(self):
        column = _FIRST_SPECIES_COLUMN
        while column <= self._sheet.max_column:
            name = self._sheet.cell(_SPECIES_ROW, column).value
            if name:
                name = str(name).strip()
                self.species.append(name)
                self._species_column[name] = column
                labels, offsets = self._scan_lifestages(column)
                self._lifestages[name] = labels
                self._offsets[name] = offsets
            column += _COLUMNS_PER_SPECIES

    def _scan_lifestages(self, species_column):
        """Lifestage labels that carry a velocity curve, with their column offsets."""
        labels, offsets = [], {}
        for offset in LIFESTAGE_OFFSETS:
            label = self._sheet.cell(_LIFESTAGE_ROW, species_column + offset - 1).value
            has_curve = self._sheet.cell(PARAMETER_ROWS["u"],
                                         species_column + offset - 1).value is not None
            if label and has_curve:
                label = str(label).strip()
                if label not in offsets:
                    labels.append(label)
                    offsets[label] = offset
        return labels, offsets

    def lifestages(self, species):
        """Lifestages available for a species, as the workbook labels them."""
        try:
            species = self.resolve_species(species)
        except KeyError:
            return []
        return list(self._lifestages.get(species, []))

    def pairs(self):
        """Yield every ``(species, lifestage)`` the database defines."""
        for species in self.species:
            for lifestage in self._lifestages[species]:
                yield species, lifestage

    @staticmethod
    def shortname(species, lifestage):
        """Four-letter code the original used in file names: ``chsp``, ``chju``, ..."""
        return (str(species).lower()[:2] + str(lifestage).lower()[:2])

    def curve(self, species, lifestage, parameter):
        """Habitat suitability curve for one species, lifestage and parameter.

        Args:
            species (str): common name, as in the workbook.
            lifestage (str): ``"spawning"``, ``"fry"``, ``"juvenile"`` or ``"adult"``.
            parameter (str): ``"u"``, ``"h"`` or ``"substrate"``.

        Returns:
            tuple: ``(x, y)`` arrays, or ``None`` when the workbook defines no curve.
        """
        if parameter not in PARAMETER_ROWS:
            raise KeyError("unknown parameter %r" % parameter)
        species = self.resolve_species(species)
        lifestage = self.resolve_lifestage(species, lifestage)
        offset = self._offsets[species][lifestage]

        base = self._species_column[species]
        x_column, y_column = base + offset - 1, base + offset
        row = PARAMETER_ROWS[parameter]

        x_values, y_values = [], []
        while row <= self._sheet.max_row:
            x = self._sheet.cell(row, x_column).value
            y = self._sheet.cell(row, y_column).value
            if x is None or y is None:
                break
            try:
                x_values.append(float(x))
                y_values.append(float(y))
            except (TypeError, ValueError):
                break
            row += 1

        if len(x_values) < 2:
            return None
        return np.asarray(x_values, dtype="float64"), np.asarray(y_values, dtype="float64")

    def cover_value(self, species, lifestage, cover_type):
        """Radius and suitability of a cover type, or None when it is not defined."""
        try:
            species = self.resolve_species(species)
            lifestage = self.resolve_lifestage(species, lifestage)
        except KeyError:
            return None
        offset = self._offsets[species][lifestage]
        base = self._species_column[species]
        row = PARAMETER_ROWS.get(cover_type)
        if row is None:
            return None
        radius = self._sheet.cell(row, base + offset - 1).value
        suitability = self._sheet.cell(row, base + offset).value
        if radius is None or suitability is None:
            return None
        return float(radius), float(suitability)

    def season_dates(self, species, lifestage):
        """Start and end of the lifestage's season, as ``(month, day)`` pairs.

        Rows 6 and 7 of the lifestage's *suitability* column, which is where
        ``cFish.get_season_dates`` read them. The year in the cell is a placeholder and is
        ignored; a season that ends before it starts wraps into the next year.

        Returns:
            tuple: ``((start_month, start_day), (end_month, end_day))``, or ``None`` when
            the workbook gives no season for this lifestage.
        """
        import datetime as dt

        try:
            species = self.resolve_species(species)
            lifestage = self.resolve_lifestage(species, lifestage)
        except KeyError:
            return None
        offset = self._offsets[species][lifestage]
        base = self._species_column[species]

        dates = []
        for key in ("start_date", "end_date"):
            value = self._sheet.cell(PARAMETER_ROWS[key], base + offset).value
            if isinstance(value, (dt.datetime, dt.date)):
                dates.append((value.month, value.day))
            else:
                return None
        return tuple(dates)

    def travel_thresholds(self, species, lifestage):
        """``{"h_min": ..., "u_max": ...}`` swimming thresholds, where defined."""
        try:
            species = self.resolve_species(species)
            lifestage = self.resolve_lifestage(species, lifestage)
        except KeyError:
            return {}
        offset = self._offsets[species][lifestage]
        base = self._species_column[species]
        thresholds = {}
        for key in ("h_min", "u_max"):
            value = self._sheet.cell(PARAMETER_ROWS[key], base + offset - 1).value
            if value is not None:
                thresholds[key] = float(value)
        return thresholds


#: Cover types the habitat database defines a curve for, in the order the original combined
#: them. ``substrate`` is a suitability *curve* on grain size; the rest are presence layers
#: whose influence spreads over a radius.
COVER_TYPES = ("substrate", "cobbles", "boulders", "plants", "wood")

#: Grain diameter range that defines each grain-based cover type, in **metres**, as
#: ``cHSI.CovHSI.define_grain_size`` had them. Converted for a condition in U.S. customary
#: units.
GRAIN_SIZE_LIMITS = {"cobbles": (0.064, 0.256), "boulders": (0.256, 100.0)}


def cover_hsi(species, lifestage, layers, profile, unit="us", fish=None, depth=None):
    """Cover habitat suitability: shelter from substrate, cobbles, boulders, plants and wood.

    Depth and velocity say whether a fish *can* be somewhere. Cover says whether it is safe
    to be there - a fry in open water at the right depth and velocity is still a meal. The
    original offered this as the *cover* option of SHArC and it is what separates a hydraulic
    habitat model from an ecohydraulic one.

    Two kinds of layer, dispatched as ``cHSI.CovHSI.call_analysis`` did:

    ``substrate``
        a grain size raster, mapped through its suitability curve like depth or velocity.
    ``cobbles``, ``boulders``, ``plants``, ``wood``
        presence layers. ``cobbles`` and ``boulders`` are cut out of a grain size raster by
        the diameter ranges in :data:`GRAIN_SIZE_LIMITS`; ``plants`` and ``wood`` are given
        directly. Each element then shelters everything within its **radius**, taken from
        the first x value of its curve, and those cells take the curve's suitability.

    The result is the cell-wise **maximum** across the types present: the best shelter
    available at a cell is what counts, not the sum of every kind of it.

    Args:
        species (str): common name; case and spacing are ignored.
        lifestage (str): lifestage; case and spacing are ignored.
        layers (dict): ``{cover type: array}``. ``cobbles`` and ``boulders`` may be omitted
            and derived from a ``substrate`` grain size raster instead.
        profile (dict): the raster profile the layers are on, for the cell size.
        unit (str): ``"us"`` or ``"si"``; the grain limits are converted accordingly.
        fish (FishDatabase): the curve database. Built by default.
        depth (numpy.ndarray): flow depth. When given, cover is cropped to cells at least as
            deep as the first point of the depth curve, as the original's
            ``crop_input_raster`` did - cover a fish cannot reach shelters nothing.

    Returns:
        tuple: ``(cover, used)`` - the suitability array with NoData where no cover applies,
        and the list of cover types that contributed. ``(None, [])`` when none did.
    """
    from . import raster

    fish = fish or FishDatabase()
    species = fish.resolve_species(species)
    lifestage = fish.resolve_lifestage(species, lifestage)
    dx, dy = raster.cell_size(profile)
    factor = 1.0 / config.FT2M if str(unit).lower() == "us" else 1.0

    reachable = None
    if depth is not None:
        depth_curve = fish.curve(species, lifestage, "h")
        h_min = float(depth_curve[0][0]) if depth_curve is not None else 0.0
        with np.errstate(invalid="ignore"):
            # `Con(h >= h_min, cover)` in the original, and Con excludes NoData: a cell the
            # 2D model never covered is not shallow, it is unknown.
            reachable = np.isfinite(depth) & (depth >= h_min)

    grain = layers.get("substrate")
    contributions, used = [], []

    for cover_type in COVER_TYPES:
        curve = fish.curve(species, lifestage, cover_type)

        if cover_type == "substrate":
            if grain is None or curve is None:
                continue
            suitability = apply_curve(grain, curve)
            if reachable is not None:
                suitability = raster.con(reachable, suitability)
            contributions.append(suitability)
            used.append(cover_type)
            continue

        # Presence layers. The curve gives a radius and the suitability inside it; where the
        # workbook has only the scalar pair, cover_value supplies it.
        value = fish.cover_value(species, lifestage, cover_type)
        if curve is not None:
            radius, suitability_value = float(curve[0][0]), float(curve[1][0])
        elif value is not None:
            radius, suitability_value = value
        else:
            continue

        present = layers.get(cover_type)
        if present is None and cover_type in GRAIN_SIZE_LIMITS and grain is not None:
            low, high = (limit * factor for limit in GRAIN_SIZE_LIMITS[cover_type])
            with np.errstate(invalid="ignore"):
                present = (grain >= low) & (grain < high)
        if present is None:
            continue

        mask = np.isfinite(present) & (np.nan_to_num(present) > 0.0) \
            if present.dtype != bool else present
        if reachable is not None:
            mask = mask & reachable
        if not mask.any():
            continue

        sheltered = raster.within_radius(mask, radius, dx, dy)
        contributions.append(raster.con(sheltered, suitability_value))
        used.append(cover_type)
        logger.info("      * cover %-10s radius %.2f, HSI %.2f, %d cell(s) sheltered",
                    cover_type, radius, suitability_value, int(sheltered.sum()))

    if not contributions:
        return None, []
    # The best shelter available wins, as CellStatistics(..., "MAXIMUM") did.
    return raster.cell_statistics(contributions, "MAXIMUM"), used


def apply_curve(array, curve):
    """Map raster values onto a habitat suitability index through a piecewise-linear curve.

    Replaces ``cHSI.HHSI.nested_con_raster_calc``. Below the first curve point the first
    suitability is held; above the last point the suitability is **zero**, not held - a
    depth or velocity beyond the curve is unsuitable, not maximally suitable, and clamping
    there would invent habitat at the highest discharges.

    Args:
        array (numpy.ndarray): depth, velocity or grain size, with ``numpy.nan`` NoData.
        curve (tuple): ``(x, y)`` from :meth:`FishDatabase.curve`.

    Returns:
        numpy.ndarray: suitability index, NoData preserved.
    """
    x_values, y_values = curve
    order = np.argsort(x_values)
    x_values, y_values = np.asarray(x_values)[order], np.asarray(y_values)[order]

    result = np.interp(np.nan_to_num(array, nan=0.0), x_values, y_values,
                       left=y_values[0], right=0.0)
    return np.where(np.isfinite(array), result, np.nan)


def read_flow_duration(path):
    """Read a flow duration workbook: ``{discharge: cumulative % exceedance}``.

    The ``2D data`` sheet of the original ``flow_duration_<shortname>.xlsx`` pairs each
    modelled discharge with the fraction of the season it is exceeded, which is the weight
    SHArea integrates over.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook["2D data"] if "2D data" in workbook.sheetnames else workbook.active

    duration = {}
    for row in range(3, sheet.max_row + 1):
        discharge = sheet.cell(row, 2).value
        exceedance = sheet.cell(row, 3).value
        if discharge is None or exceedance is None:
            continue
        try:
            duration[float(discharge)] = float(exceedance)
        except (TypeError, ValueError):
            continue
    return duration


class SHArC:
    """Habitat suitability analysis for one condition.

    Args:
        condition (Condition or str): the condition, or its name.
        unit (str): ``"us"`` or ``"si"``; must match the condition's rasters.
        combine_method (str): ``"geometric_mean"`` or ``"product"``.
        fish (FishDatabase): the curve database. Defaults to the packaged ``Fish.xlsx``.
        threshold (float): cHSI above which a cell counts as usable habitat.

    Attributes:
        error (bool): True when at least one discharge could not be processed.
    """

    def __init__(self, condition, unit="us", combine_method="geometric_mean", fish=None,
                 threshold=0.4):
        self.condition = condition if isinstance(condition, Condition) \
            else Condition(condition)
        self.unit = str(unit).lower()
        if combine_method not in COMBINE_METHODS:
            raise ValueError("combine_method must be one of %s" % (COMBINE_METHODS,))
        self.combine_method = combine_method
        self.threshold = float(threshold)
        self.fish = fish or FishDatabase()
        self.logger = logger
        self.error = False
        #: Cover types that contributed to the last run, for the result summary.
        self._cover_used = set()

        self._depth = {}
        self._velocity = {}
        for name in self.condition.all_depth_rasters():
            discharge = self.condition.discharge_of(name)
            if discharge is not None:
                self._depth[discharge] = self.condition.path(name)
        for name in self.condition.all_velocity_rasters():
            discharge = self.condition.discharge_of(name)
            if discharge is not None:
                self._velocity[discharge] = self.condition.path(name)

    @property
    def discharges(self):
        """Discharges with both a depth and a velocity raster, ascending."""
        return sorted(set(self._depth) & set(self._velocity))

    # ------------------------------------------------------------------- rasters

    def composite_hsi(self, species, lifestage, discharge, reference=None, cover=None,
                      cover_layers=None):
        """Composite habitat suitability at one discharge.

        Args:
            species (str): common name; case and spacing are ignored.
            lifestage (str): lifestage; case and spacing are ignored.
            discharge (float): the discharge to evaluate.
            reference (dict): profile every raster is aligned onto.
            cover (numpy.ndarray): a ready-made cover suitability array.
            cover_layers (dict): cover *inputs*, from which one is built with
                :func:`cover_hsi`. Ignored when ``cover`` is given.

        Returns:
            tuple: ``(chsi, profile)``, or ``(None, None)`` when the curves are missing.
        """
        depth_curve = self.fish.curve(species, lifestage, "h")
        velocity_curve = self.fish.curve(species, lifestage, "u")
        if depth_curve is None or velocity_curve is None:
            self.logger.info("      * no depth/velocity curve for %s %s", species, lifestage)
            return None, None

        depth, depth_profile = raster.read(self._depth[discharge])
        velocity, velocity_profile = raster.read(self._velocity[discharge])
        reference = reference or depth_profile
        depth = raster.align(depth, depth_profile, reference)
        velocity = raster.align(velocity, velocity_profile, reference)

        dsi = apply_curve(depth, depth_curve)
        vsi = apply_curve(velocity, velocity_curve)

        if cover is None and cover_layers:
            cover, used = cover_hsi(species, lifestage, cover_layers, reference,
                                    unit=self.unit, fish=self.fish, depth=depth)
            self._cover_used.update(used)

        with np.errstate(invalid="ignore"):
            if cover is not None:
                product = dsi * vsi * cover
                chsi = np.cbrt(product) if self.combine_method == "geometric_mean" \
                    else product
            else:
                product = dsi * vsi
                chsi = np.sqrt(product) if self.combine_method == "geometric_mean" \
                    else product

        # Habitat exists only in the wetted area. Two-argument con, so dry cells are NoData
        # rather than a suitability of zero, which would otherwise dilute the mean cHSI.
        return raster.con(np.nan_to_num(depth) > 0.0, chsi), reference

    def usable_area(self, chsi, profile, weighted=False):
        """Area where cHSI exceeds the threshold, optionally weighted by the mean cHSI."""
        dx, dy = raster.cell_size(profile)
        relevant = raster.con(chsi > self.threshold, chsi)
        mask = np.isfinite(relevant)
        area = float(mask.sum() * dx * dy)
        if weighted and mask.any():
            area *= float(np.nanmean(relevant))
        return area

    # ----------------------------------------------------------------------- run

    def cover_layers(self, extra=None):
        """Cover layers available in the condition, on its own grid.

        Looks for the rasters the cover analysis can use: the grain size raster named by
        ``input_definitions.inp`` for ``substrate`` (and, through it, cobbles and boulders),
        and ``plants.tif`` / ``wood.tif`` beside the condition.

        Args:
            extra (dict): ``{cover type: path or array}`` to add or override.

        Returns:
            dict: ``{cover type: array}``, empty when the condition has no cover data.
        """
        found = {}
        if self.condition.exists(self.condition.grain_raster):
            found["substrate"] = self.condition.path(self.condition.grain_raster)
        for cover_type in ("cobbles", "boulders", "plants", "wood"):
            if self.condition.exists(cover_type):
                found[cover_type] = self.condition.path(cover_type)
        found.update(extra or {})

        layers = {}
        for cover_type, source in found.items():
            if isinstance(source, str):
                array, _profile = raster.read(source)
                layers[cover_type] = array
            elif source is not None:
                layers[cover_type] = source
        return layers

    def run(self, species, lifestage, discharges=None, output_dir=None,
            weighted=False, write_rasters=True, flow_duration=None, cover=False):
        """Map habitat suitability across discharges and compute SHArea.

        Args:
            species (str): common name from the fish database.
            lifestage (str): lifestage from the fish database.
            discharges (list): discharges to evaluate. Defaults to all available.
            output_dir (str): where the cHSI rasters go.
            weighted (bool): weight usable area by the mean cHSI, as the original's
                *apply weighting* option did.
            write_rasters (bool): write one ``csi_<code><Q>.tif`` per discharge.
            flow_duration (str or dict): a flow duration workbook, or a
                ``{discharge: % exceedance}`` mapping. Without it SHArea is not computed.
            cover (bool or dict): include a cover habitat suitability index. ``True`` uses
                whatever :meth:`cover_layers` finds in the condition; a dict supplies or
                overrides layers. See :func:`cover_hsi`.

        Returns:
            dict: ``per_discharge`` rows, ``sharea`` when a flow duration was given, and
            the paths written.
        """
        species = self.fish.resolve_species(species)
        lifestage = self.fish.resolve_lifestage(species, lifestage)
        code = self.fish.shortname(species, lifestage)
        discharges = sorted(discharges) if discharges else self.discharges
        discharges = [q for q in discharges if q in self._depth and q in self._velocity]
        if not discharges:
            raise ValueError("no discharge has both a depth and a velocity raster")

        output_dir = output_dir or os.path.join(config.dir_output("SHArC"),
                                                self.condition.name)
        if write_rasters:
            os.makedirs(output_dir, exist_ok=True)

        reference = raster.profile_of(self._depth[discharges[0]])

        # Cover does not change with discharge except through the depth crop, so it is
        # built once and cropped per discharge inside composite_hsi.
        self._cover_used = set()
        cover_layers = None
        if cover is not False and cover is not None:
            cover_layers = self.cover_layers(cover if isinstance(cover, dict) else None)
            if not cover_layers:
                self.logger.info("      * no cover layer found in condition %r - running "
                                 "without cover", self.condition.name)

        rows = []
        for discharge in discharges:
            chsi, profile = self.composite_hsi(species, lifestage, discharge, reference,
                                               cover_layers=cover_layers)
            if chsi is None:
                self.error = True
                continue
            area = self.usable_area(chsi, profile, weighted=weighted)
            row = {"discharge": discharge, "usable_area": area}
            with np.errstate(invalid="ignore"):
                finite = chsi[np.isfinite(chsi)]
            row["mean_chsi"] = float(finite.mean()) if finite.size else 0.0
            rows.append(row)

            if write_rasters:
                path = os.path.join(output_dir, "csi_%s%06d.tif" % (code, discharge))
                raster.write(path, chsi, profile)
                row["raster"] = path

        result = {
            "condition": self.condition.name,
            "species": species,
            "lifestage": lifestage,
            "shortname": code,
            "combine_method": self.combine_method,
            "threshold": self.threshold,
            "weighted": weighted,
            "cover": sorted(self._cover_used),
            "per_discharge": rows,
            "area_unit": config.area_unit(self.unit),
            "discharge_unit": config.unit_labels(self.unit)["q"],
        }
        if write_rasters:
            result["output_dir"] = output_dir

        duration = self._resolve_flow_duration(flow_duration, code)
        if duration:
            result["sharea"] = self.sharea(rows, duration)
            result["flow_duration"] = duration
        return result

    def _resolve_flow_duration(self, flow_duration, code):
        """Accept a mapping, a workbook path, or find the workbook for this species code."""
        if isinstance(flow_duration, dict):
            return flow_duration
        if isinstance(flow_duration, str):
            return read_flow_duration(flow_duration)

        pattern = os.path.join(config.dir_flows(), self.condition.name,
                               "flow_duration_%s.xlsx" % code)
        matches = sorted(glob.glob(pattern))
        if not matches:
            self.logger.info("      * no flow duration workbook for %s - SHArea not "
                             "computed", code)
            return None
        return read_flow_duration(matches[0])

    @staticmethod
    def sharea(rows, flow_duration):
        """Integrate usable area over the flow duration curve.

        Reproduces the ``(E_i - E_{i-1}) / 100 * F_i`` column of the original workbook: a
        Riemann sum of usable area over cumulative exceedance probability, so habitat that
        exists only at a rare discharge contributes little.

        Args:
            rows (list): ``{"discharge": ..., "usable_area": ...}`` entries.
            flow_duration (dict): ``{discharge: cumulative % exceedance}``.

        Returns:
            float: seasonal habitat area, in the area unit of the usable areas.
        """
        paired = [(row["discharge"], row["usable_area"]) for row in rows
                  if row["discharge"] in flow_duration]
        if not paired:
            return 0.0
        # Ascending exceedance: the rarest (highest) discharge first, as the workbook had it.
        paired.sort(key=lambda item: flow_duration[item[0]])

        total = 0.0
        previous = 0.0
        for discharge, area in paired:
            exceedance = flow_duration[discharge]
            total += (exceedance - previous) / 100.0 * area
            previous = exceedance
        return total
