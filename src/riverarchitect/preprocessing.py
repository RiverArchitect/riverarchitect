"""Preparing a condition: the terrain products every other module depends on.

The open-source replacement for the ArcGIS ``GetStarted`` module. Nothing here is an
analysis in its own right; it produces the derived rasters that lifespan mapping, habitat
suitability and recruitment all read:

* :func:`detrended_dem` - elevation above the local thalweg, which is what makes an
  elevation comparable between the upstream and downstream ends of a reach;
* :func:`water_level_elevation`, :func:`interpolated_depth` and :func:`depth_to_water_table`
  - a continuous water surface extrapolated from the wetted area, and the depth and
  depth-to-groundwater rasters derived from it;
* :func:`morphological_units` - a depth and velocity classification into pools, riffles,
  runs and the rest, after Wyrick and Pasternack (2014);
* :func:`write_input_definitions` - the ``input_definitions.inp`` a condition needs;
* :func:`align_condition` - put every raster of a condition on one grid.

Relation to the original
------------------------
The original did the interpolation steps by converting rasters to point shapefiles, running
``SpatialJoin_analysis`` with ``match_option="CLOSEST"``, and converting back with
``PointToRaster_conversion``. That round trip through vector data was a way to get a
nearest-neighbour interpolation out of arcpy; here it is
:func:`riverarchitect.raster.nearest_neighbour` directly, with
:func:`riverarchitect.raster.idw` and :func:`riverarchitect.raster.kriging` available as
alternatives the original could not offer.
"""

import logging
import os

import numpy as np

from . import config, raster

__all__ = ["detrended_dem", "water_level_elevation", "interpolated_depth",
           "depth_to_water_table", "morphological_units", "MorphologicalUnits",
           "write_input_definitions", "align_condition", "build_product",
           "INTERPOLATION_METHODS", "PRODUCTS"]

logger = logging.getLogger("riverarchitect")

#: Interpolation methods accepted where a surface is extrapolated from wetted cells.
INTERPOLATION_METHODS = ("nearest", "idw", "kriging")


def _interpolate(points, values, profile, method="nearest", **kwargs):
    """Dispatch to the requested interpolator in :mod:`riverarchitect.raster`."""
    if method == "nearest":
        return raster.nearest_neighbour(points, values, profile)
    if method == "idw":
        return raster.idw(points, values, profile, **kwargs)
    if method == "kriging":
        return raster.kriging(points, values, profile, **kwargs)
    raise ValueError("method must be one of %s" % (INTERPOLATION_METHODS,))


def _sample_points(array, profile, step=1):
    """Points and values of the finite cells of an array."""
    points, values = raster.raster_to_points(array, profile, step=step)
    finite = np.isfinite(values)
    return points[finite], values[finite]


# ------------------------------------------------------------------- detrended DEM

def detrended_dem(dem_path, depth_path, output_path=None, method="nearest", step=1):
    """Elevation above the local thalweg.

    A raw DEM cannot be compared along a reach: 3 ft above the bed means something different
    where the bed is 10 ft higher. Detrending removes the downstream slope by subtracting the
    elevation of the nearest wetted cell - the thalweg at the discharge given.

    Args:
        dem_path (str): the digital elevation model.
        depth_path (str): a flow depth raster; its wetted cells define the thalweg. Use a
            low, in-channel discharge.
        output_path (str): where to write the result. Optional.
        method (str): ``"nearest"`` (the original's behaviour), ``"idw"`` or ``"kriging"``.
        step (int): sample every n-th thalweg cell. Raise it on very large rasters.

    Returns:
        tuple: ``(detrended, profile)``.
    """
    dem, profile = raster.read(dem_path)
    depth, depth_profile = raster.read(depth_path)
    depth = raster.align(depth, depth_profile, profile)

    # Thalweg elevation: the bed where it is wet. con with two arguments, so dry cells are
    # NoData and are not sampled as if they were at elevation zero.
    thalweg = raster.con(np.nan_to_num(depth) > 0.0, dem)
    points, values = _sample_points(thalweg, profile, step=step)
    if points.size == 0:
        raise ValueError("no wetted cell in %s - cannot locate a thalweg" % depth_path)

    logger.info("   >> detrending against %d thalweg cells (%s)", len(values), method)
    surface = _interpolate(points, values, profile, method=method)
    detrended = dem - surface

    if output_path:
        raster.write(output_path, detrended, profile)
    return detrended, profile


# ------------------------------------------------------------------- water levels

def water_level_elevation(dem_path, depth_path, output_path=None, method="nearest",
                          step=1):
    """Extrapolate a continuous water surface from the wetted area.

    In the wetted area the water surface is ``dem + depth``. Outside it there is no
    modelled water surface at all, so it is interpolated - which is what makes a
    depth-to-groundwater raster possible on dry land.

    Returns:
        tuple: ``(wle, profile)``.
    """
    dem, profile = raster.read(dem_path)
    depth, depth_profile = raster.read(depth_path)
    depth = raster.align(depth, depth_profile, profile)

    wetted = np.nan_to_num(depth) > 0.0
    surface = raster.con(wetted, dem + depth)
    points, values = _sample_points(surface, profile, step=step)
    if points.size == 0:
        raise ValueError("no wetted cell in %s - cannot build a water surface" % depth_path)

    logger.info("   >> interpolating the water surface from %d wetted cells (%s)",
                len(values), method)
    wle = _interpolate(points, values, profile, method=method)

    if output_path:
        raster.write(output_path, wle, profile)
    return wle, profile


def interpolated_depth(dem_path, depth_path, output_path=None, wle=None, **kwargs):
    """Flow depth extended beyond the modelled wetted area.

    ``wle - dem``, kept where positive. Used where a 2D model covers less than the area an
    analysis needs.

    Returns:
        tuple: ``(depth, profile)``.
    """
    dem, profile = raster.read(dem_path)
    if wle is None:
        wle, profile = water_level_elevation(dem_path, depth_path, **kwargs)

    depth = wle - dem
    depth = raster.con(depth > 0.0, depth)

    if output_path:
        raster.write(output_path, depth, profile)
    return depth, profile


def depth_to_water_table(dem_path, depth_path, output_path=None, wle=None, **kwargs):
    """Depth from the ground surface down to the water table.

    ``dem - wle``: positive on dry land above the water surface, which is the range
    vegetation-planting features are keyed to. Cells below the water surface are negative,
    and are kept rather than clipped - a planting feature needs to know it is under water.

    Returns:
        tuple: ``(d2w, profile)``.
    """
    dem, profile = raster.read(dem_path)
    if wle is None:
        wle, profile = water_level_elevation(dem_path, depth_path, **kwargs)

    d2w = dem - wle

    if output_path:
        raster.write(output_path, d2w, profile)
    return d2w, profile


# -------------------------------------------------------------- morphological units

class MorphologicalUnits:
    """Depth and velocity thresholds that define each morphological unit.

    Read from a ``morphological_units.xlsx`` in the original layout: column D the unit name,
    E its raster code, F and G the depth range, H and I the velocity range. The workbook is
    in **SI**, so the thresholds are converted when the condition is in U.S. customary units.

    Args:
        path (str): the workbook. Defaults to the one shipped with the package.
        unit (str): unit system of the rasters the thresholds will be applied to.
    """

    _FIRST_ROW = 6
    _LAST_ROW = 44

    def __init__(self, path=None, unit="us"):
        import warnings

        import openpyxl

        self.path = path or os.path.join(config.templates_dir(), "morphological_units.xlsx")
        if not os.path.isfile(self.path):
            raise FileNotFoundError("no morphological unit table at %s" % self.path)
        self.unit = str(unit).lower()
        # The workbook is metric; 1 m = 1/0.3048 ft, and the same factor applies to m/s.
        self.factor = 1.0 / config.FT2M if self.unit == "us" else 1.0

        self.units = {}
        with warnings.catch_warnings():
            # openpyxl drops the workbook's data-validation extension on read and says so.
            # It does not affect the thresholds, and there is nothing to act on.
            warnings.simplefilter("ignore", UserWarning)
            sheet = openpyxl.load_workbook(self.path, data_only=True).active
        for row in range(self._FIRST_ROW, self._LAST_ROW + 1):
            name = sheet.cell(row, 4).value
            code = sheet.cell(row, 5).value
            if not name or code is None or str(name).strip().lower() in ("none", "mu type"):
                continue
            bounds = [sheet.cell(row, column).value for column in (6, 7, 8, 9)]
            if any(value is None for value in bounds):
                continue
            h_min, h_max, u_min, u_max = (float(value) * self.factor for value in bounds)
            self.units[str(name).strip()] = {
                "code": int(code), "h_min": h_min, "h_max": h_max,
                "u_min": u_min, "u_max": u_max,
            }

    def codes(self):
        """``{unit name: raster code}``, as :mod:`riverarchitect.lifespan` expects it."""
        return {name.lower(): entry["code"] for name, entry in self.units.items()}

    def __len__(self):
        return len(self.units)

    def __repr__(self):
        return "MorphologicalUnits(%d units, unit=%r)" % (len(self.units), self.unit)


def morphological_units(depth_path, velocity_path, output_path=None, table=None,
                        unit="us"):
    """Classify the wetted area into morphological units by depth and velocity.

    After Wyrick and Pasternack (2014). A cell takes the unit whose depth *and* velocity
    range it falls into; where ranges overlap the highest code wins, which is what the
    original's ``CellStatistics(..., "MAXIMUM")`` did.

    Args:
        depth_path (str): flow depth raster, usually at baseflow.
        velocity_path (str): flow velocity raster at the same discharge.
        output_path (str): where to write the result. Optional.
        table (MorphologicalUnits): the threshold table. Built by default.
        unit (str): unit system of the rasters.

    Returns:
        tuple: ``(mu, profile, table)``.
    """
    table = table or MorphologicalUnits(unit=unit)
    depth, profile = raster.read(depth_path)
    velocity, velocity_profile = raster.read(velocity_path)
    velocity = raster.align(velocity, velocity_profile, profile)

    wet = np.nan_to_num(depth) > 0.0
    layers = []
    for name, entry in table.units.items():
        with np.errstate(invalid="ignore"):
            selected = (wet
                        & (depth >= entry["h_min"]) & (depth < entry["h_max"])
                        & (velocity >= entry["u_min"]) & (velocity < entry["u_max"]))
        if selected.any():
            layers.append(raster.con(selected, float(entry["code"])))
            logger.info("   >> %-22s code %-3d %d cell(s)", name, entry["code"],
                        int(selected.sum()))

    if not layers:
        raise ValueError("no cell falls into any morphological unit - check the units of "
                         "the depth and velocity rasters against the table")

    mu = raster.cell_statistics(layers, "MAXIMUM")
    if output_path:
        raster.write(output_path, mu, profile)
    return mu, profile, table


# ------------------------------------------------------------------ condition setup

def write_input_definitions(condition_dir, return_periods=None, discharges=None,
                            path=None, **rasters):
    """Write the ``input_definitions.inp`` that names a condition's rasters.

    Args:
        condition_dir (str): the condition folder.
        return_periods (list): flood return period per discharge, in years.
        discharges (list): discharges the return periods belong to. Defaults to every
            ``h<Q>.tif`` on disk, ascending.
        path (str): output path. Defaults to ``<condition_dir>/input_definitions.inp``.
        **rasters: override a default raster name, e.g. ``grain_raster="d50"``.

    Returns:
        str: the path written.
    """
    from .condition import Condition

    condition_dir = os.path.abspath(condition_dir)
    path = path or os.path.join(condition_dir, "input_definitions.inp")

    if discharges is None:
        discharges = sorted(
            q for q in (Condition.discharge_of(name) for name in os.listdir(condition_dir)
                        if name.lower().startswith("h") and name.lower().endswith(".tif"))
            if q is not None)

    if return_periods and len(return_periods) != len(discharges):
        raise ValueError("%d return periods for %d discharges - they must correspond"
                         % (len(return_periods), len(discharges)))

    names = {"grain_raster": "dmean", "detrended_raster": "dem_detrend",
             "d2w_raster": "d2w", "mu_raster": "mu", "dem_raster": "dem"}
    names.update({key: value for key, value in rasters.items() if value})

    depth = ", ".join("h%06d.tif" % q for q in discharges)
    velocity = ", ".join("u%06d.tif" % q for q in discharges)
    periods = ", ".join(str(value) for value in (return_periods or []))

    lines = [
        "# RASTER META DATA - ONLY MODIFY VALUES BETWEEN '=' AND '#'",
        "# Written by riverarchitect.preprocessing.write_input_definitions",
        "#" + "-" * 87,
        "Return periods = %s #[Comma separated LIST] defines lifespans" % periods,
        "#",
        "# RASTER NAMES",
        "#" + "-" * 87,
        "Flow depth (h) = %s #[Comma separated LIST]" % depth,
        "Flow velocity (u) = %s #[Comma separated LIST]" % velocity,
        "Grain sizes (D mean) = %s #[STRING]" % names["grain_raster"],
        "Detrended DEM = %s #[STRING]" % names["detrended_raster"],
        "Depth to groundwater table (d2w) = %s #[STRING]" % names["d2w_raster"],
        "Morphological units (mu) = %s #[STRING]" % names["mu_raster"],
        "DEM = %s #[STRING]" % names["dem_raster"],
        "DEM of differences = fill, scour #[Comma separated LIST]",
        "",
    ]
    with open(path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(lines))
    logger.info("   >> wrote %s (%d discharges)", path, len(discharges))
    return path


def align_condition(condition_dir, reference=None, output_dir=None, pattern="*.tif"):
    """Resample every raster of a condition onto one grid.

    Rasters assembled from different preprocessing chains routinely differ in extent and
    cell size - in the sample condition the DEM of difference is on a 5 ft grid while
    everything else is on 3 ft. Analyses call :func:`riverarchitect.raster.align` per
    operand, so this is a convenience rather than a requirement; it is worth doing once when
    a condition is going to be used repeatedly.

    Args:
        condition_dir (str): the condition folder.
        reference (str): raster whose grid to adopt. Defaults to ``dem.tif``, else the first.
        output_dir (str): where to write. Defaults to writing in place.
        pattern (str): which rasters to align.

    Returns:
        dict: ``{path: "aligned" | "unchanged"}``.
    """
    paths = raster.list_rasters(condition_dir, pattern)
    if not paths:
        raise FileNotFoundError("no rasters matching %s in %s" % (pattern, condition_dir))

    if reference is None:
        preferred = os.path.join(condition_dir, "dem.tif")
        reference = preferred if os.path.isfile(preferred) else paths[0]
    reference_profile = raster.profile_of(reference)
    output_dir = output_dir or condition_dir
    os.makedirs(output_dir, exist_ok=True)

    results = {}
    for path in paths:
        array, profile = raster.read(path)
        same_grid = (profile["width"] == reference_profile["width"]
                     and profile["height"] == reference_profile["height"]
                     and profile["transform"] == reference_profile["transform"])
        target = os.path.join(output_dir, os.path.basename(path))
        if same_grid and os.path.abspath(target) == os.path.abspath(path):
            results[path] = "unchanged"
            continue
        aligned = array if same_grid else raster.align(array, profile, reference_profile)
        raster.write(target, aligned, reference_profile)
        results[path] = "unchanged" if same_grid else "aligned"
    logger.info("   >> aligned %d of %d raster(s) onto %s",
                sum(1 for value in results.values() if value == "aligned"), len(results),
                os.path.basename(reference))
    return results


#: The products :func:`build_product` can make, as ``(label, key)``. Both front ends render
#: this list, so they cannot drift apart.
PRODUCTS = (
    ("detrended DEM", "detrended"),
    ("water surface, depth and depth to water table", "water"),
    ("morphological units", "mu"),
    ("input_definitions.inp", "inp"),
    ("align every raster onto one grid", "align"),
)

#: One-line explanation of each product, shown in the interface.
PRODUCT_NOTES = {
    "detrended": "Elevation above the local thalweg, so elevations are comparable along "
                 "the reach. Writes dem_detrend.tif.",
    "water": "Extrapolates the water surface from the wetted area, then writes wle.tif, "
             "h_interp.tif and d2w.tif.",
    "mu": "Classifies the wetted area into pools, riffles, runs and the rest from depth "
          "and velocity. Writes mu.tif.",
    "inp": "Writes the input_definitions.inp that names the condition's rasters. Add flood "
           "return periods afterwards for lifespan mapping.",
    "align": "Resamples every raster of the condition onto the DEM's grid. Optional - the "
             "analyses align per operand anyway.",
}


def build_product(condition_name, key, discharge=None, method="nearest", unit="us",
                  output_dir=None):
    """Build one named product for a condition, and report what was written.

    The single entry point both the Qt and the tkinter interface call, so neither has to
    know how a product is assembled and the two cannot drift apart.

    Args:
        condition_name (str): the condition.
        key (str): one of the keys in :data:`PRODUCTS`.
        discharge (float): reference discharge, for the products that need one.
        method (str): interpolation method, see :data:`INTERPOLATION_METHODS`.
        unit (str): unit system of the rasters.
        output_dir (str): where to write. Defaults to the condition folder.

    Returns:
        list: lines describing what was written.
    """
    from .condition import Condition

    condition = Condition(condition_name)
    target = output_dir or condition.directory
    os.makedirs(target, exist_ok=True)
    dem = condition.path(condition.dem_raster)
    depth = condition.path("h%06d.tif" % discharge) if discharge else None
    lines = []

    if key == "detrended":
        path = os.path.join(target, "dem_detrend.tif")
        detrended_dem(dem, depth, path, method=method)
        lines.append("wrote %s" % path)
    elif key == "water":
        wle_path = os.path.join(target, "wle.tif")
        wle, _profile = water_level_elevation(dem, depth, wle_path, method=method)
        lines.append("wrote %s" % wle_path)
        for maker, filename in ((interpolated_depth, "h_interp.tif"),
                                (depth_to_water_table, "d2w.tif")):
            path = os.path.join(target, filename)
            maker(dem, depth, path, wle=wle)
            lines.append("wrote %s" % path)
    elif key == "mu":
        velocity = condition.path("u%06d.tif" % discharge)
        path = os.path.join(target, "mu.tif")
        _mu, _profile, table = morphological_units(depth, velocity, path, unit=unit)
        lines.append("wrote %s (%d unit types)" % (path, len(table)))
    elif key == "inp":
        path = write_input_definitions(condition.directory)
        lines.append("wrote %s" % path)
        lines.append("")
        lines.append("Return periods are left empty. Fill them in for lifespan mapping:")
        lines.append("one value per discharge, in years.")
    elif key == "align":
        results = align_condition(condition.directory, output_dir=output_dir)
        changed = sum(1 for value in results.values() if value == "aligned")
        lines.append("aligned %d of %d raster(s)" % (changed, len(results)))
    else:
        raise ValueError("unknown product %r" % key)
    return lines
