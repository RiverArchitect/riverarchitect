"""Lifespan and design mapping for river restoration features.

The open-source replacement for the ArcGIS ``LifespanDesign`` module. It answers two
questions per cell, for a chosen feature:

**Lifespan** - how many years is this feature expected to survive here? Each modelled
discharge carries a flood return period. For every criterion the feature has a threshold
for; the analysis finds the lowest return period at which the threshold is exceeded, which
is when the feature fails. Cells that survive every modelled flood stay NoData: their
lifespan is longer than the largest modelled event and cannot be quantified from the data.

**Design** - how big does the feature have to be to reach a target lifespan? For grain-based
features that is the stable grain size at the design flood.

.. note::
   The streamwood design map of the original, which sized a stable log from water depth and
   Froude number rather than from a Shields criterion, is **not** implemented. Streamwood
   still gets a lifespan map; it simply gets no design map, and
   :meth:`LifespanDesign.run_feature` omits ``design_raster`` from its result rather than
   writing a wrong one.

Threshold values come from :data:`FEATURES`, which reproduces the defaults of the original
``threshold_values.xlsx``. Load a customised workbook with :func:`load_threshold_workbook`,
or build a :class:`Feature` directly.

Relation to the original
------------------------
This implements the analysis hierarchy as documented in the wiki (see
``docs/wiki/LifespanDesign.md``) rather than transcribing the arcpy control flow, which
threaded results through a mutable ``raster_dict`` and a chain of ``Con()`` calls whose
behaviour depended on which rasters happened to exist. The semantics are the documented ones:

* every hydraulic criterion produces a *failure return period* raster, and the lifespan is
  their cell-wise **minimum** - the first flood that breaks the feature;
* the spatial criteria (depth to water table, detrended DEM, morphological units,
  topographic change, terrain slope) are **masks** applied to that result;
* two-argument ``con`` is used throughout, so the false branch is NoData, never zero.
"""

import logging
import os

import numpy as np

from . import config, raster, shear
from .condition import Condition

__all__ = ["Feature", "FEATURES", "LifespanDesign", "load_threshold_workbook",
           "feature_groups"]

logger = logging.getLogger("riverarchitect")

#: Relative grain density (ratio of sediment and water density).
RHO_RATIO = shear.RHO_RATIO
#: Default Manning's n in s/m^(1/3). Divided by 1.49 for U.S. customary units.
MANNING_N = 0.0473934
#: Gravitational acceleration in m/s^2.
G_SI = shear.G_SI


class Feature:
    """Threshold values for one restoration feature.

    Every threshold is optional. A criterion whose threshold is ``None``, or whose input
    raster the condition does not have, is skipped - which is how the original behaved and
    why a sparsely defined feature still produces a map.

    Args:
        fid (str): short feature id, used in output file names (``lf_<fid>.tif``).
        name (str): human-readable name.
        group (str): feature group, for the interface.
        tau_cr (float): critical dimensionless bed shear stress.
        safety_factor (float): safety factor for the mobile-grain analysis. When set, the
            mobile-grain analysis runs *instead of* the dimensionless shear-stress one.
        h_max (float): water depth above which the feature fails, in the condition's length
            unit.
        u_max (float): flow velocity above which the feature fails.
        froude_max (float): Froude number above which the feature fails.
        d2w_min, d2w_max (float): depth-to-water-table range the feature needs.
        det_min, det_max (float): detrended-DEM range the feature is restricted to.
        grain_max (float): largest admissible grain size, for fine-sediment features.
        terrain_slope (float): slope above which nature-based engineering is required.
        scour_rate, fill_rate (float): topographic change thresholds.
        inverse_tcd (bool): True when the feature is relevant *below* those rates rather
            than above them.
        design_frequency (float): target lifespan in years for the design map.
        mu_avoid, mu_relevant (list): morphological units to exclude or restrict to.
        mu_method (int): 0 to avoid ``mu_avoid``, 1 to keep only ``mu_relevant``.
        lifespan_mapping, design_mapping (bool): which maps the feature supports.
    """

    def __init__(self, fid, name, group="", tau_cr=None, safety_factor=None,
                 h_max=None, u_max=None, froude_max=None,
                 d2w_min=None, d2w_max=None, det_min=None, det_max=None,
                 grain_max=None, terrain_slope=None,
                 scour_rate=None, fill_rate=None, inverse_tcd=False,
                 design_frequency=None, mu_avoid=None, mu_relevant=None, mu_method=0,
                 lifespan_mapping=True, design_mapping=False):
        self.fid = fid
        self.name = name
        self.group = group
        self.tau_cr = tau_cr
        self.safety_factor = safety_factor
        self.h_max = h_max
        self.u_max = u_max
        self.froude_max = froude_max
        self.d2w_min = d2w_min
        self.d2w_max = d2w_max
        self.det_min = det_min
        self.det_max = det_max
        self.grain_max = grain_max
        self.terrain_slope = terrain_slope
        self.scour_rate = scour_rate
        self.fill_rate = fill_rate
        self.inverse_tcd = inverse_tcd
        self.design_frequency = design_frequency
        self.mu_avoid = list(mu_avoid or [])
        self.mu_relevant = list(mu_relevant or [])
        self.mu_method = mu_method
        self.lifespan_mapping = lifespan_mapping
        self.design_mapping = design_mapping

    def __repr__(self):
        return "Feature(%r, %r)" % (self.fid, self.name)


def _f(*features):
    return {feature.fid: feature for feature in features}


#: Default features and thresholds, reproducing ``LifespanDesign/.templates/threshold_values.xlsx``
#: of the original. Values are in U.S. customary units, as that workbook was.
FEATURES = _f(
    # --- terraforming ---------------------------------------------------------------
    Feature("backwt", "Backwater", "Terraforming", tau_cr=0.047, u_max=0.1,
            design_frequency=5, fill_rate=0.3, scour_rate=0.3,
            mu_relevant=["agriplain", "backswamp", "mining pit", "pond", "slackwater"],
            mu_method=1, design_mapping=True),
    Feature("widen", "Widen", "Terraforming", det_min=17, det_max=25,
            mu_relevant=["bank", "floodplain", "high floodplain", "island-floodplain",
                         "island high floodplain", "in-channel bar", "lateral bar",
                         "levee", "spur dike", "terrace"],
            mu_method=1, lifespan_mapping=False, design_mapping=True),
    Feature("grade", "Grading", "Terraforming", tau_cr=0.047, d2w_min=7, d2w_max=12,
            scour_rate=0.3, inverse_tcd=True, mu_avoid=["bedrock", "hillside"]),
    Feature("sideca", "Side cavities", "Terraforming", fill_rate=1.0, inverse_tcd=True,
            mu_relevant=["bank", "cutbank", "in-channel bar", "lateral bar", "spur dike",
                         "tailings"],
            mu_method=1, lifespan_mapping=False, design_mapping=True),
    Feature("sidech", "Side channels", "Terraforming", tau_cr=0.047, fill_rate=1.0,
            inverse_tcd=True),

    # --- vegetation plantings -------------------------------------------------------
    Feature("Generic", "Generic planting", "Vegetation plantings", tau_cr=0.06),
    Feature("box", "Box Elder", "Vegetation plantings", tau_cr=0.06,
            d2w_min=1, d2w_max=7, h_max=1),
    Feature("cot", "Cottonwood", "Vegetation plantings", d2w_min=1, d2w_max=7,
            h_max=2.1, u_max=3),
    Feature("whi", "White Alder", "Vegetation plantings", tau_cr=0.06, d2w_min=1, d2w_max=7),
    Feature("wil", "Willow", "Vegetation plantings", tau_cr=0.1, d2w_min=1, d2w_max=7,
            h_max=2.1),

    # --- established vegetation -----------------------------------------------------
    Feature("Generic_est", "Generic (established)", "Established vegetation", tau_cr=0.06),
    Feature("Box_est", "Box Elder (established)", "Established vegetation", tau_cr=0.06,
            d2w_min=1, d2w_max=7, h_max=1),
    Feature("cot_est", "Cottonwood (established)", "Established vegetation",
            d2w_min=1, d2w_max=7, h_max=2.1, u_max=3),
    Feature("Whi_est", "White Alder (established)", "Established vegetation", tau_cr=0.06,
            d2w_min=1, d2w_max=7),
    Feature("Wil_est", "Willow (established)", "Established vegetation", tau_cr=0.1,
            d2w_min=1, d2w_max=7, h_max=2.1),

    # --- other nature-based engineering ---------------------------------------------
    Feature("wood", "Streamwood", "Nature-based engineering", h_max=3.34, froude_max=1,
            design_frequency=10, design_mapping=True),
    Feature("rocks", "Angular boulders", "Nature-based engineering", tau_cr=0.047,
            safety_factor=1.3, design_frequency=20, scour_rate=3, design_mapping=True),
    Feature("bio", "Other nature-based eng.", "Nature-based engineering", d2w_max=12,
            terrain_slope=0.2, design_mapping=True),

    # --- connectivity and gravel augmentation ---------------------------------------
    Feature("gravin", "Gravel: In", "Connectivity", tau_cr=0.047, design_frequency=10,
            mu_relevant=["chute", "fast glide", "flood runner", "bedrock", "in-channel bar",
                         "lateral bar", "medial bar", "pool", "riffle", "riffle transition",
                         "run", "slackwater", "slow glide", "swale", "tailings"],
            mu_method=1, design_mapping=True),
    Feature("gravou", "Gravel: Out", "Connectivity", tau_cr=0.047, design_frequency=1,
            scour_rate=3,
            mu_relevant=["agriplain", "backswamp", "bank", "cutbank", "flood runner",
                         "floodplain", "high floodplain", "hillside",
                         "island high floodplain", "island-floodplain", "in-channel bar",
                         "lateral bar", "levee", "medial bar", "mining pit", "point bar",
                         "pond", "spur dike", "tailings", "terrace"],
            mu_method=1, design_mapping=True),
    Feature("fines", "Incorporation of fine sediment", "Connectivity", tau_cr=0.03,
            d2w_min=1, d2w_max=10, grain_max=0.0066667, fill_rate=3.36, scour_rate=3,
            design_mapping=True),
)


def feature_groups(features=None):
    """Group features by :attr:`Feature.group`, preserving definition order.

    Returns:
        dict: group name -> list of :class:`Feature`.
    """
    groups = {}
    for feature in (features or FEATURES).values():
        groups.setdefault(feature.group or "Other", []).append(feature)
    return groups


def load_threshold_workbook(path):
    """Read a ``threshold_values.xlsx`` in the original layout.

    Lets a project keep using a workbook it has already calibrated. Row 5 holds the feature
    id and each following row one threshold, exactly as the original expected.

    Args:
        path (str): path to the workbook.

    Returns:
        dict: feature id -> :class:`Feature`.
    """
    import openpyxl

    rows = {
        6: "bed_shear", 7: "tau_cr", 8: "d2w_min", 9: "d2w_max", 10: "det_min",
        11: "det_max", 12: "h_max", 13: "u_max", 14: "froude_max", 15: "grain_max",
        16: "design_frequency", 17: "mu_avoid", 18: "mu_relevant", 19: "mu_method",
        20: "safety_factor", 21: "terrain_slope", 22: "inverse_tcd", 23: "fill_rate",
        24: "scour_rate", 25: "lifespan_mapping", 26: "design_mapping",
    }
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.active

    features = {}
    for column in range(5, sheet.max_column + 1):
        fid = sheet.cell(5, column).value
        if not fid:
            continue
        kwargs = {"name": sheet.cell(4, column).value or str(fid)}
        for row, attribute in rows.items():
            value = sheet.cell(row, column).value
            if value is None or attribute == "bed_shear":
                continue
            if attribute in ("mu_avoid", "mu_relevant"):
                items = [item.strip() for item in str(value).split(",") if item.strip()]
                kwargs[attribute] = [item for item in items if item.lower() != "na"]
            elif attribute in ("inverse_tcd", "lifespan_mapping", "design_mapping"):
                kwargs[attribute] = bool(value)
            elif attribute == "mu_method":
                kwargs[attribute] = int(value)
            else:
                kwargs[attribute] = float(value)
        features[str(fid)] = Feature(str(fid), **kwargs)
    return features


class LifespanDesign:
    """Run lifespan and design mapping for one condition.

    Args:
        condition (Condition or str): the condition, or its name.
        unit (str): ``"us"`` or ``"si"``; must match the condition's rasters.
        manning_n (float): Manning's n in s/m^(1/3); converted for U.S. customary units.
        features (dict): feature id -> :class:`Feature`. Defaults to :data:`FEATURES`.

    Attributes:
        error (bool): True when at least one feature could not be mapped.
    """

    def __init__(self, condition, unit="us", manning_n=MANNING_N, features=None):
        self.condition = condition if isinstance(condition, Condition) \
            else Condition(condition)
        self.unit = str(unit).lower()
        self.features = features or FEATURES
        self.error = False
        self.logger = logger

        self.n = manning_n / 1.49 if self.unit == "us" else manning_n  # s/ft^(1/3)
        self.g = shear.gravity_of(self.unit)

        self._reference = None                 # profile every raster is aligned onto
        self._taux_cache = {}                  # discharge token -> theta84
        self._shear_diagnostics = {}           # discharge token -> (h_over_ks, regime)
        self._diagnostics_written = set()      # output dirs already holding them

    # ----------------------------------------------------------------------- inputs

    def _read(self, name, aligned=True):
        """Read a raster of the condition, aligned onto the reference grid, or None."""
        if not self.condition.exists(name):
            return None
        array, profile = raster.read(self.condition.path(name))
        if self._reference is None:
            self._reference = profile
            return array
        return raster.align(array, profile, self._reference) if aligned else array

    def _set_reference(self):
        """Pick the grid everything is resampled onto: the grain raster, else the DEM."""
        for name in (self.condition.grain_raster, self.condition.dem_raster):
            if self.condition.exists(name):
                _, self._reference = raster.read(self.condition.path(name))
                return self._reference
        for _period, depth_path, _u in self.condition.hydraulic_pairs():
            self._reference = raster.profile_of(depth_path)
            return self._reference
        raise FileNotFoundError("condition %r has no usable raster to define the grid.\n\n%s"
                                % (self.condition.name, self.condition.describe()))

    def _hydraulics(self):
        """Yield ``(return_period, token, depth, velocity)`` on the reference grid.

        ``token`` is the discharge in file-name form, e.g. ``"000550"`` - it keys the
        per-discharge shear diagnostics and their output file names, the same way
        :mod:`riverarchitect.recruitment` names its own.
        """
        for period, depth_path, velocity_path in self.condition.hydraulic_pairs():
            discharge = self.condition.discharge_of(depth_path)
            token = self.condition.token_for(discharge) if discharge is not None \
                else os.path.splitext(os.path.basename(depth_path))[0]
            depth, depth_profile = raster.read(depth_path)
            velocity, velocity_profile = raster.read(velocity_path)
            depth = raster.align(depth, depth_profile, self._reference)
            velocity = raster.align(velocity, velocity_profile, self._reference)
            # Dry cells must be NoData, not zero: h**(1/3) in the denominator of the
            # mobile-grain formula otherwise divides by zero across the whole dry bed.
            yield period, token, np.where(depth > 0, depth, np.nan), velocity

    # -------------------------------------------------------------------- criteria

    @staticmethod
    def _first_failure(failures):
        """Cell-wise earliest failure return period. Replaces ``compare_raster_set``."""
        if not failures:
            return None
        return raster.cell_statistics(failures, "MINIMUM")

    def critical_grain_size(self, depth, velocity, tau_cr, safety_factor=1.0):
        """Grain diameter that just starts to move, in the condition's length unit."""
        with np.errstate(invalid="ignore", divide="ignore"):
            return ((velocity * self.n) ** 2
                    / ((RHO_RATIO - 1.0) * tau_cr * depth ** (1.0 / 3.0))
                    / safety_factor)

    def shields_stress(self, depth, velocity, grain, token=None):
        """Dimensionless bed shear stress ``theta84``, referenced to ``D84 = 2.2 * Dmean``.

        Replaces the Keulegan-only expression of the original (arcpy counterpart:
        ``analyse_taux``) with the regime-aware closure of :mod:`riverarchitect.shear`,
        which stays physically meaningful in the shallow cells the logarithmic law cannot
        describe. When ``token`` is given, the relative submergence and regime rasters of
        that discharge are kept for :meth:`write_shear_diagnostics`.
        """
        if token is not None and token in self._taux_cache:
            return self._taux_cache[token]
        result = shear.calculate_taux(velocity, depth, shear.d84_of(grain),
                                      gravity=self.g)
        if token is not None:
            # theta84 does not depend on the feature, so one computation per discharge
            # serves every taux feature; the result is kept for writing once.
            self._taux_cache[token] = result.theta84
            self._shear_diagnostics[token] = result
            self.logger.info("      * taux %s: %s", token,
                             ", ".join("%s %d" % (label, count) for label, count
                                       in shear.regime_summary(result.regime).items()))
            if not result.regime.any():
                self.logger.warning(
                    "      * taux %s is NoData everywhere: the grain raster %r may not "
                    "share units or extent with the hydraulic rasters.", token,
                    self.condition.grain_raster)
        return result.theta84

    def _hydraulic_lifespan(self, feature, grain):
        """Earliest failure return period across every hydraulic criterion."""
        per_criterion = []
        depth_fail, velocity_fail, froude_fail, grain_fail = [], [], [], []

        for period, token, depth, velocity in self._hydraulics():
            if feature.h_max is not None:
                depth_fail.append(raster.con(depth >= feature.h_max, period))
            if feature.u_max is not None:
                velocity_fail.append(raster.con(velocity >= feature.u_max, period))
            if feature.froude_max is not None:
                with np.errstate(invalid="ignore", divide="ignore"):
                    froude = velocity / np.sqrt(self.g * depth)
                froude_fail.append(raster.con(froude >= feature.froude_max, period))
            if feature.tau_cr is not None and grain is not None:
                if feature.safety_factor is not None:
                    # Mobile grains: the feature fails once the flow can move a grain of
                    # the size actually present.
                    mobile = self.critical_grain_size(depth, velocity, feature.tau_cr,
                                                      feature.safety_factor)
                    grain_fail.append(raster.con(mobile >= grain, period))
                else:
                    taux = self.shields_stress(depth, velocity, grain, token=token)
                    grain_fail.append(raster.con(taux >= feature.tau_cr, period))

        for failures in (depth_fail, velocity_fail, froude_fail, grain_fail):
            combined = self._first_failure(failures)
            if combined is not None:
                per_criterion.append(combined)

        return self._first_failure(per_criterion)

    def _spatial_mask(self, feature, grain):
        """Boolean mask of cells where the feature is admissible at all, or None."""
        masks = []

        if feature.d2w_min is not None or feature.d2w_max is not None:
            d2w = self._read(self.condition.d2w_raster)
            if d2w is not None:
                low = -np.inf if feature.d2w_min is None else feature.d2w_min
                high = np.inf if feature.d2w_max is None else feature.d2w_max
                masks.append((d2w >= low) & (d2w <= high))

        if feature.det_min is not None or feature.det_max is not None:
            detrended = self._read(self.condition.detrended_raster)
            if detrended is not None:
                low = -np.inf if feature.det_min is None else feature.det_min
                high = np.inf if feature.det_max is None else feature.det_max
                masks.append((detrended >= low) & (detrended <= high))

        if feature.grain_max is not None and grain is not None:
            masks.append(grain <= feature.grain_max)

        if feature.terrain_slope is not None:
            dem = self._read(self.condition.dem_raster)
            if dem is not None:
                dx, dy = raster.cell_size(self._reference)
                slope = raster.slope(dem, dx, dy, units="PERCENT") / 100.0
                masks.append(slope >= feature.terrain_slope)

        tcd = self._topographic_change_mask(feature)
        if tcd is not None:
            masks.append(tcd)

        mu = self._morphological_unit_mask(feature)
        if mu is not None:
            masks.append(mu)

        if not masks:
            return None
        combined = masks[0]
        for mask in masks[1:]:
            combined = combined & mask
        return combined

    def _topographic_change_mask(self, feature):
        """Where scour or fill exceeds the feature's rate - or stays below it, if inverse."""
        rates = []
        for name, threshold in (("scour", feature.scour_rate), ("fill", feature.fill_rate)):
            if threshold is None:
                continue
            array = self._read(name)
            if array is None:
                continue
            rates.append(array < threshold if feature.inverse_tcd else array >= threshold)
        if not rates:
            return None
        combined = rates[0]
        for rate in rates[1:]:
            # Either kind of change makes the feature relevant; both must stay low to make
            # an inverse-relevance feature (grading, side cavities) admissible.
            combined = (combined & rate) if feature.inverse_tcd else (combined | rate)
        return combined

    def _morphological_unit_mask(self, feature):
        """Restrict to, or exclude, listed morphological units.

        Needs a ``morphological_units.xlsx``-style code mapping to be meaningful. Without
        one the criterion is skipped rather than guessed at, and that is logged.
        """
        wanted = feature.mu_relevant if feature.mu_method == 1 else feature.mu_avoid
        if not wanted:
            return None
        if not self.condition.exists(self.condition.mu_raster):
            return None
        codes = self._mu_codes()
        if not codes:
            self.logger.info("      * no morphological-unit code table - skipping the MU "
                             "criterion for %s", feature.fid)
            return None
        mu = self._read(self.condition.mu_raster)
        if mu is None:
            return None
        selected = np.zeros(mu.shape, dtype=bool)
        unknown = []
        for unit in wanted:
            code = codes.get(str(unit).strip().lower())
            if code is None:
                unknown.append(str(unit))
                continue
            selected |= (mu == code)
        if unknown:
            # Naming the units that did not resolve matters: the original swallowed this in
            # a bare except and dropped the whole criterion without saying so.
            self.logger.info("      * %s: no code for morphological unit(s) %s - not part "
                             "of the criterion", feature.fid, ", ".join(unknown))
        return selected if feature.mu_method == 1 else ~selected

    def _mu_codes(self):
        """Morphological unit name -> raster code.

        A ``morphological_units.xlsx`` beside the condition wins, so a project that has
        recoded its units is honoured. Otherwise the table packaged with River Architect is
        used, which is where the original read its codes from as well
        (``cParameters.MU.read_mus``); falling back to it is what makes the criterion apply
        at all on a condition that ships no workbook of its own.
        """
        if hasattr(self, "_mu_code_cache"):
            return self._mu_code_cache
        self._mu_code_cache = {}

        for candidate in ("morphological_units.xlsx", "mu.xlsx"):
            path = os.path.join(self.condition.directory, candidate)
            if not os.path.isfile(path):
                continue
            try:
                from .preprocessing import MorphologicalUnits
                self._mu_code_cache = MorphologicalUnits(path, unit=self.unit).codes()
            except Exception as exc:  # a malformed table must not stop the analysis
                self.logger.info("      * could not read %s (%s)", candidate, exc)
            break

        if not self._mu_code_cache:
            try:
                from .preprocessing import MorphologicalUnits
                self._mu_code_cache = MorphologicalUnits(unit=self.unit).codes()
            except Exception as exc:
                self.logger.info("      * could not read the packaged morphological unit "
                                 "table (%s)", exc)
        return self._mu_code_cache

    # ------------------------------------------------------------------------- run

    def run_feature(self, feature, output_dir):
        """Map one feature. Returns a summary dict, or None when nothing could be mapped."""
        if isinstance(feature, str):
            feature = self.features[feature]
        if self._reference is None:
            self._set_reference()

        grain = self._read(self.condition.grain_raster)
        lifespan = self._hydraulic_lifespan(feature, grain)

        if lifespan is None:
            # No hydraulic criterion applies. The feature is then admissible wherever its
            # spatial criteria allow, for as long as the condition can demonstrate.
            mask = self._spatial_mask(feature, grain)
            if mask is None:
                self.logger.info("   >> %s: no applicable criteria - skipped", feature.fid)
                return None
            lifespan = raster.con(mask, float(self.condition.max_lifespan))
        else:
            mask = self._spatial_mask(feature, grain)
            if mask is not None:
                lifespan = raster.con(mask, lifespan)

        os.makedirs(output_dir, exist_ok=True)
        result = {"feature": feature.fid, "name": feature.name}

        if feature.lifespan_mapping:
            path = os.path.join(output_dir, "lf_%s.tif" % feature.fid)
            raster.write(path, lifespan, self._reference)
            result["lifespan_raster"] = path

        if feature.design_mapping:
            design = self._design_raster(feature, lifespan, grain)
            if design is not None:
                path = os.path.join(output_dir, "ds_%s.tif" % feature.fid)
                raster.write(path, design, self._reference)
                result["design_raster"] = path

        self.write_shear_diagnostics(output_dir)

        dx, dy = raster.cell_size(self._reference)
        mapped = np.isfinite(lifespan)
        result["area"] = float(mapped.sum() * dx * dy)
        result["area_unit"] = config.area_unit(self.unit)
        if mapped.any():
            result["min_lifespan"] = float(np.nanmin(lifespan))
            result["max_lifespan"] = float(np.nanmax(lifespan))
            result["median_lifespan"] = float(np.nanmedian(lifespan))
        return result

    def write_shear_diagnostics(self, output_dir):
        """Write the bed shear stress rasters of each discharge beside the lifespan maps.

        Four per discharge - ``ts`` (Shields stress), ``tb`` (``u*^2``), ``hks`` (relative
        submergence) and ``regime`` (which resistance closure applied) - named exactly as
        :func:`riverarchitect.preprocessing.bed_shear_stress` names them in the condition
        folder. Idempotent per output directory and per run.
        """
        if not self._shear_diagnostics or output_dir in self._diagnostics_written:
            return
        from .preprocessing import write_shear_rasters

        for token, result in self._shear_diagnostics.items():
            write_shear_rasters(result, self._reference, output_dir, token)
        self._diagnostics_written.add(output_dir)

    def _design_raster(self, feature, lifespan, grain):
        """Stable grain size, or stable log diameter, at the feature's design flood."""
        target = feature.design_frequency
        if target is None or feature.tau_cr is None or grain is None:
            return None

        chosen = None
        for period, _token, depth, velocity in self._hydraulics():
            if period >= target:
                chosen = (depth, velocity)
                break
        if chosen is None:
            self.logger.info("      * no discharge reaches the %s-year design flood for %s",
                             target, feature.fid)
            return None

        depth, velocity = chosen
        design = self.critical_grain_size(depth, velocity, feature.tau_cr,
                                          feature.safety_factor or 1.0)
        if self.unit == "us":
            design = design * 12.0                     # ft -> in
        return raster.con(np.isfinite(lifespan), design)

    def run(self, feature_ids=None, output_dir=None):
        """Map several features.

        Args:
            feature_ids (list): feature ids; defaults to every feature that supports
                lifespan mapping.
            output_dir (str): where the rasters go. Defaults to
                ``Output/LifespanDesign/<condition>``.

        Returns:
            list: one summary dict per mapped feature.
        """
        output_dir = output_dir or os.path.join(
            config.dir_output("LifespanDesign"), self.condition.name)
        if feature_ids is None:
            feature_ids = [fid for fid, feature in self.features.items()
                           if feature.lifespan_mapping]

        self._set_reference()
        if not any(True for _pair in self.condition.hydraulic_pairs()):
            # Without hydraulics every feature would silently map nothing; say why.
            self.logger.warning("   >> no paired depth and velocity rasters - only "
                                "spatial criteria can apply.\n%s",
                                self.condition.describe())
        results = []
        for fid in feature_ids:
            feature = self.features.get(fid)
            if feature is None:
                self.logger.info("   >> unknown feature %r - skipped", fid)
                self.error = True
                continue
            self.logger.info("   >> mapping %s (%s)", feature.name, feature.fid)
            try:
                result = self.run_feature(feature, output_dir)
            except Exception as exc:
                self.logger.error("   >> %s failed: %s", feature.fid, exc)
                self.error = True
                continue
            if result:
                results.append(result)
        return results
