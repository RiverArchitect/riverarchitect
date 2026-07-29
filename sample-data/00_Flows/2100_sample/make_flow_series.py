"""Build the synthetic daily flow record ``flow_series_2020.csv`` beside this script.

The sample condition ships flow *duration* curves but no dated record, and the recruitment
box model needs one - bed preparation, recession and scour are all about *when* a flow
happened, not just which flows are possible. This shapes a single water year whose flow
duration matches the condition's own ``Q data`` sheet, so the **discharges are the reach's
own** and only their ordering in time is invented.

That ordering is a plausible Mediterranean-climate water year: a wet winter, three storms, a
spring snowmelt freshet and a long summer recession. It is *not* a gauge record, so treat
recruitment results on this condition as a demonstration of the method rather than as a
finding about the reach.

Regenerate with::

    python sample-data/00_Flows/2100_sample/make_flow_series.py \\
        sample-data/00_Flows/2100_sample/flow_series_2020.csv
"""
import datetime as dt
import os
import sys

import numpy as np
import openpyxl

SOURCE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                      "flow_duration_chsp.xlsx")
YEAR = 2020


def duration_curve(path):
    """(% exceedance ascending, discharge) from the 'Q data' sheet."""
    sheet = openpyxl.load_workbook(path, data_only=True)["Q data"]
    pairs = []
    for row in range(3, sheet.max_row + 1):
        q, exceedance = sheet.cell(row, 1).value, sheet.cell(row, 2).value
        if q is None or exceedance is None:
            continue
        pairs.append((float(exceedance), float(q)))
    pairs.sort()
    return np.array([p[0] for p in pairs]), np.array([p[1] for p in pairs])


def shape(year):
    """Exceedance probability per day of a Mediterranean-climate water year.

    Water year starts 1 October. Low exceedance (rare, high flow) in the winter rain and
    spring snowmelt months, high exceedance (common, low flow) through the dry summer.
    """
    start = dt.date(year - 1, 10, 1)
    days = (dt.date(year, 9, 30) - start).days + 1
    t = np.arange(days)

    # Seasonal baseline: a smooth annual cycle peaking in late February.
    peak = (dt.date(year, 2, 20) - start).days
    seasonal = 50.0 - 42.0 * np.cos(2.0 * np.pi * (t - peak) / days)

    # Three winter storms and a snowmelt freshet, each a short excursion to a rare flow.
    events = [((year, 1, 8), 12, 0.05), ((year, 2, 18), 9, 0.3),
              ((year, 3, 21), 7, 1.5), ((year, 5, 20), 25, 6.0)]
    for (y, m, d), width, low in events:
        centre = (dt.date(y, m, d) - start).days
        bell = np.exp(-0.5 * ((t - centre) / (width / 2.5)) ** 2)
        seasonal = seasonal * (1.0 - bell) + low * bell

    return start, days, np.clip(seasonal, 0.02, 99.9)


def main():
    exceedance, discharge = duration_curve(SOURCE)
    start, days, daily_exceedance = shape(YEAR)
    # The duration curve is exceedance-ascending, so discharge descends; np.interp needs an
    # ascending x, which it already is.
    q = np.interp(daily_exceedance, exceedance, discharge)

    target = sys.argv[1] if len(sys.argv) > 1 else "flow_series_2020.csv"
    os.makedirs(os.path.dirname(os.path.abspath(target)), exist_ok=True)
    with open(target, "w", encoding="utf-8") as handle:
        handle.write("Date,Mean daily\n")
        for offset in range(days):
            handle.write("%s,%.1f\n" % (start + dt.timedelta(days=offset), q[offset]))

    print("wrote %s: %d days, %.0f - %.0f cfs, median %.0f cfs"
          % (target, days, q.min(), q.max(), np.median(q)))


if __name__ == "__main__":
    main()
